import os
import pyperclip
import jieba
from time import time
from openpyxl import styles, load_workbook
from auto_moudle import AutoMationModule
from lib import FunWrap, log, sleep
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
jieba.load_userdict("./input/userdict.txt")


class ChuanQi(AutoMationModule):
    def __init__(self):
        # 初始化父类方法
        #  super(ChuanQi, self).__init__()
        #  self.jgy = JianGuoYun("1403341393@qq.com", "abhdwrkfxxrnhnyf")
        self.set_script_message = None
        # 显示脚本状态的信号
        self.init_var()

    def init_driver(self, port=None, browser_name="360安全浏览器"):
        options = Options()
        browser_path = self.get_browser_path(browser_name)
        options.binary_location = browser_path
        if port:
            options.add_experimental_option(
                "debuggerAddress", f"127.0.0.1:{port}")
        self.driver = webdriver.Chrome(
            "./input/chromedriver.exe", options=options)
        self.dl(f"连接{browser_name}成功!")

    def get_browser_path(self, browser_name):
        import win32com.client
        try:
            shell = win32com.client.Dispatch("WScript.Shell")
            path = os.path.join(os.path.expanduser(r'~'), "Desktop")
            for dirpath, dirnames, filenames in os.walk(path):
                for filename in filenames:
                    if browser_name in filename:
                        path = dirpath + "\\" + filename
            shortcut = shell.CreateShortCut(path).TargetPath
            return shortcut
        except Exception as e:
            print(e)
            raise Exception(f"{browser_name}路径查找失败")

    def init_var(self):
        # 脚本执行状态
        self.status = "初始化"
        self.thread_status = "初始化"
        # 执行参数
        self.params = {}
        # 自定义参数----
        self.point_list = [
            {
                "name": "搜索人数",
                "point": (1020, 250),
            },
            {
                "name": "支付转化率",
                "point": (1000, 430),
            },
            {
                "name": "在线商品数",
                "point": (1170, 450),
            },
            {
                "name": "商城点击占比",
                "point": (1160, 490),
            },
            {
                "name": "空白处",
                "point": (700, 820),
            },
            {
                "name": "关键词输入",
                "point": (1200, 560),
            },
        ]
        self.black_list = self.get_black_list()

    def pretreatment(self, task_name=""):
        ''' 预处理函数(每个函数开始之前执行) '''
        if "stop" in self.status:
            self.exit()
            self.set_status("停止脚本")
            self.dl("停止线程")
            raise Exception("停止线程")
        elif "暂停" in self.thread_status:
            self.dl("暂停线程")
            while "暂停" in self.thread_status:
                self.wait(1)
        if task_name != "":
            log.info(task_name + " 开始")
            self.set_status(task_name)

    def set_status(self, status_text):
        ''' 设置脚本状态 '''
        self.status = status_text
        if self.set_script_message:
            self.set_script_message.emit(self.index, status_text)

    @FunWrap("获取关键词", True)
    def get_black_list(self):
        book = load_workbook("./out/标题裂变产品.xlsx")
        print(1)
        sheet = book['黑名单']
        _list = []
        for row in range(1, sheet.max_row + 1):
            for col in range(1, 2):
                value = sheet.cell(row, col).value
                if not value:
                    continue
                _list.append(value)
        log.info(_list)
        return _list

    @FunWrap("读取产品列表", True)
    def read_product_list(self):
        self.product_list = []
        book = load_workbook("./out/标题裂变产品.xlsx")
        sheet = book['产品']
        for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=2):
            temp_list = []
            for cell in row:
                if not cell.value:
                    break
                temp_list.append(cell.value)
            if temp_list:
                self.product_list.append(temp_list)

    def wait(self, sec, is_while=False):
        if is_while:
            for _ in range(sec):
                self.wait(1)
                sec -= 1
                self.set_status(f"等待: {sec}")
        else:
            sleep(sec)

    ################  复制标题  ###################

    def get_number(self, temp_str):
        return int("".join(list(filter(str.isdigit, temp_str))))

    @FunWrap("获取同款标题", True)
    def get_title_list(self, link):
        self.driver.implicitly_wait(1)
        self.driver.get(link)
        timeout = time() + 10
        elem = None
        while timeout > time():
            elem = self.driver.find_elements_by_xpath(
                '//*[@class="layui-tab-item t1 layui-show"]//*[@class="lxSimilarSale"]')
            if elem:
                break
            sleep(0.5)
        if not elem:
            return None

        def sort_a(e):
            return self.get_number(e.get_attribute("textContent"))
        elem_list = self.driver.find_elements_by_xpath(
            '//*[@class="layui-tab-item t1 layui-show"]//*[@class="lxSimilarSale"]')
        elem_list.sort(key=sort_a, reverse=True)
        title_list = []
        for i, e in enumerate(elem_list):
            if self.get_number(e.get_attribute("textContent")) <= 0:
                break
            elif i >= 60:
                break
            title_list.append(self.driver.execute_script(
                "return arguments[0].parentNode", e).get_attribute("title"))
        title_list = list(set(title_list))
        if self.params.get("checkbox_cut_enable", True):
            for i, title in enumerate(title_list):
                title_list[i] = " ".join(jieba.cut(title))
        return title_list

    @FunWrap("保存同款标题", True)
    def save_title(self, sheet_name, title_list):
        book = load_workbook("./out/标题裂变产品.xlsx")  # 打开exce文件
        if sheet_name not in book.sheetnames:
            book.create_sheet(sheet_name)
        sheet = book[sheet_name]
        horizontal_heander_list = [("标题", 60, 'A'), ("产品词", 10, 'B'), (
            "蓝海词", 20, 'C'), ("词根", 10, 'D'), ("裂变标题", 20, 'E'), ("", 20, "F")]
        for i, horizontal_heander in enumerate(horizontal_heander_list):
            sheet.cell(1, 1 + i).value = horizontal_heander[0]
            sheet.column_dimensions[horizontal_heander[2]
                                    ].width = horizontal_heander[1]
            sheet.cell(1, 1 + i).alignment = styles.Alignment(
                horizontal='center', vertical='center')
        for i, title in enumerate(title_list):
            sheet.cell(i+2, 1).value = title
        book.save("./out/标题裂变产品.xlsx")

    def get_title_save_to_excel(self):
        for i in range(int(self.params.get("input_product_index", 0)), len(self.product_list)):
            title_list = self.get_title_list(self.product_list[i][1])
            if title_list:
                self.dl("获取同款标题成功")
                self.save_title(self.product_list[i][0], title_list)
            else:
                self.dl("获取同款标题失败")

    ################  拆分词根  ###################

    @FunWrap("读取关键词", True)
    def read_keyword_list(self, sheet_name):
        keyword_list = []
        book = load_workbook("./out/标题裂变产品.xlsx")  # 打开exce文件
        sheet = book[sheet_name]
        for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=1):
            for cell in row:
                if not cell.value:
                    continue
                if cell.value:
                    keyword_list.append(cell.value)
        return keyword_list

    def auto_group_words(self, word_list, word_len: int):
        new_word_list = []
        last_word = ""
        is_mod = False
        for word in word_list:
            if word.isspace():
                continue
            try:
                if last_word == "":
                    raise Exception("上一个单词为空!")
                if len(word) != 1:
                    raise Exception("单词长度不等于1")
                if len(last_word) <= word_len:
                    print(f"当前组成的词语:{last_word + word}")
                    is_mod = True
            except Exception as e:
                pass
            if is_mod:
                new_word_list[-1] = last_word + word
                is_mod = False
            else:
                new_word_list.append(word)
            last_word = word
        return new_word_list

    @FunWrap("词根分析", True)
    def get_root_word_list(self, keyword_list):
        root_word_list = []
        for keyword in keyword_list:
            temp_root_word_list = jieba.cut(keyword)
            temp_root_word_list = self.auto_group_words(temp_root_word_list, 1)
            temp_root_word_list = self.auto_group_words(
                temp_root_word_list, 99)
            for root_word in temp_root_word_list:
                if root_word in root_word_list:
                    continue
                root_word_list.append(root_word)
        #  root_word_list = list(set(root_word_list))
        return root_word_list

    @FunWrap("保存词根", True)
    def save_root_word(self, sheet_name, root_word_list):
        book = load_workbook("./out/标题裂变产品.xlsx")  # 打开exce文件
        sheet = book[sheet_name]
        for row in sheet.iter_rows(min_row=2, min_col=4, max_row=sheet.max_row, max_col=4):
            for cell in row:
                cell.value = ""
        for i, root_word in enumerate(root_word_list):
            sheet.cell(i+2, 4).value = root_word
        book.save("./out/标题裂变产品.xlsx")

    def get_root_word_save_to_excel(self):
        ("获取同款标题保存到excel", True)
        for i in range(int(self.params.get("input_product_index", 0)), len(self.product_list)):
            keyword_list = self.read_keyword_list(self.product_list[i][0])
            root_word_list = self.get_root_word_list(keyword_list)
            self.save_root_word(self.product_list[i][0], root_word_list)

    ################  批量查词  ###################

    @FunWrap("读取产品词", True)
    def read_product_word_list(self, sheet_name):
        product_word_list = []
        book = load_workbook("./out/标题裂变产品.xlsx")  # 打开exce文件
        sheet = book[sheet_name]  # 打开工作表
        for row in sheet.iter_rows(min_row=2, min_col=2, max_row=sheet.max_row, max_col=2):
            for cell in row:
                if not cell.value:
                    continue
                product_word_list.append(cell.value)
        return product_word_list

    @FunWrap("保存查词数据", True)
    def save_data(self, sheet_name, data_list):
        book = load_workbook("./out/标题裂变产品.xlsx")
        sheet = book[sheet_name]
        start_index = self.find_max_col(sheet["F"])
        if sheet['F1'].value:
            data_list = data_list[1:]
        for i, data in enumerate(data_list):
            if data[0] in self.black_list:
                continue
            if data[0] == "":
                continue
            for j, value in enumerate(data):
                if value.isdigit():
                    value = int(value)
                if j == 5:
                    if i == 0 and data[0] != "关键词":
                        sheet.cell(start_index + i, 6 + j).value = "搜索"
                        sheet.cell(start_index + i, 6 + j).hyperlink = value
                    else:
                        sheet.cell(start_index + i, 6 + j).value = value
                elif j == 6:
                    if i == 0 and data[0] != "关键词":
                        sheet.cell(start_index + i, 6 + j).value = "分析"
                        sheet.cell(start_index + i, 6 + j).hyperlink = value
                    else:
                        sheet.cell(start_index + i, 6 + j).value = value
                else:
                    sheet.cell(start_index + i, 6 + j).value = value
        sheet.auto_filter.ref = f"F1:L{sheet.max_row}"
        sheet.auto_filter.add_filter_column(1, [])
        sheet.auto_filter.add_sort_condition(f"F2:L{sheet.max_row}")
        book.save("./out/标题裂变产品.xlsx")

    @FunWrap("添加黑名单", True)
    def add_black_list(self, list2):
        book = load_workbook("./out/标题裂变产品.xlsx")
        sheet = book['黑名单']
        self.del_empty_row_by_list(sheet)
        keyword_list = [li[0] for li in list2]
        for keyword in keyword_list:
            if keyword in self.black_list:
                continue
            sheet.append([keyword])
        book.save("./out/标题裂变产品.xlsx")
        self.black_list = self.get_black_list()

    def find_max_col(self, col_list):
        col_list = list(col_list)
        col_list.reverse()
        for cell in col_list:
            if cell.value:
                return cell.row
        return 1

    @FunWrap("设置固定窗口", True)
    def set_fix_window(self):
        self.dll.setWindowSize(self.hwnd, 1254, 854)

    @FunWrap("", True)
    def click_point(self, point_name, **option):
        for value in self.point_list:
            if value["name"] == point_name:
                x, y = value["point"]
        self.click(x, y, **option)

    @FunWrap("关闭窗口", True)
    def close(self):
        self.click_point("空白处")
        result = self.find_image_by_dll("关闭.bmp", is_click=True)
        if not result:
            self.dl("关闭")
            return False
        for _ in range(10):
            result = self.find_image_by_dll("批量查词.bmp", find_time=1)
            if result:
                break
        if not result:
            self.dl("未找到批量查词")
            return False
        return True

    @FunWrap("批量查词", True)
    def pi_liang_cha_ci(self, keyword):
        result = self.find_image_by_dll("批量查词.bmp", is_click=True)
        if not result:
            self.dl("未点击批量查词")
            return False
        for _ in range(10):
            result = self.find_image_by_dll("批量查询.bmp", find_time=1)
            if result:
                break
        if not result:
            self.dl("未找到批量查询")
            return False
        # 输入数据
        self.click_point("关键词输入")
        sleep(0.1)
        for _ in range(50):
            self.dll.KeyPressChar("back")
            sleep(0.01)
        # 获取5个词语
        data = keyword
        for value in data:
            sleep(0.3)
            self.input_text(value)
            self.dll.KeyPressChar("enter")
        result = self.find_image_by_dll("批量查询.bmp", is_click=True, find_time=1)
        if not result:
            self.dl("未点击批量查询")
            return False
        for _ in range(120):
            self.dl("查找复制数据中")
            result = self.find_image_by_dll("复制数据.bmp", find_time=1, sim=0.7)
            if result:
                break
        if not result:
            self.dl("未跳转表格页面")
            return False
        return True

    @FunWrap("一键筛选", False)
    def shuai_xuan(self):
        result = self.find_image_by_dll("快捷筛选.bmp", is_click=True)
        if not result:
            self.dl("未点击快捷筛选")
            return False
        for _ in range(10):
            result = self.find_image_by_dll("一键筛选.bmp", find_time=1)
            if result:
                break
        if not result:
            self.dl("未找到一键筛选")
            return False
        # 输入数据
        sleep(0.3)
        self.click_point("搜索人数")
        self.input_text(self.params.get("input_search_number", "70"))
        sleep(0.25)
        self.click_point("支付转化率")
        self.input_text(self.params.get("input_pay", "5"))
        self.click_point("在线商品数")
        self.input_text(self.params.get("input_online_product", "500"))
        sleep(0.5)
        self.click_point("商城点击占比")
        self.input_text(self.params.get("input_gdp", "50"))
        sleep(0.25)
        result = self.find_image_by_dll("一键筛选.bmp", is_click=True)
        if not result:
            self.dl("一键筛选")
            return False
        for _ in range(10):
            result = self.find_image_by_dll("复制数据.bmp", find_time=1)
            if result:
                break
        if not result:
            self.dl("未跳转表格页面")
            return False
        return True

    @FunWrap("复制数据", True)
    def copy_all_data(self):
        result = self.find_image_by_dll("复制数据.bmp", is_click=True)
        if not result:
            self.dl("未点击复制数据")
            return False
        result = self.find_image_by_dll(
            "复制全部数据.bmp", is_click=True, find_time=10)
        if not result:
            self.dl("未点击复制全部数据")
            return False
        sleep(3)
        data = pyperclip.paste()
        pyperclip.copy("")
        if isinstance(data, str):
            self.dl("返回数据列表")
            return data.replace("\r", "")
        self.dl("空空")
        return False

    def get_data_list(self, data):
        data_list = []
        for value in data.split("\n"):
            temp_data = value.split("|")
            temp_list = []
            # 筛选特定关键字
            for index, temp in enumerate(temp_data):
                if index in [0, 1, 7, 8, 9]:
                    temp_list.append(temp)
            if temp_list:
                data_list.append(temp_list)
        for i, data in enumerate(data_list):
            if data[0] == "关键词":
                data_list[i] += ["淘宝搜索", "趋势分析"]
            else:
                data_list[i] += [f'https://s.taobao.com/search?q={data[0]}',
                                 f"https://sycm.taobao.com/mc/mq/search_analyze?keyword={data[0]}"]
        return data_list

    def pi_liang_cha_ci_run(self):
        self.index = int(self.params.get("index", 0))
        self.hwnd = int(self.params.get("hwnd", 0))
        # 绑定窗口
        self.bind_window()
        try:
            self.set_fix_window()
            for i in range(int(self.params.get("input_product_index", 0)), len(self.product_list)):
                product_word_list = self.read_product_word_list(
                    self.product_list[i][0])
                self.dl(f"当前产品:{self.product_list[i][0]}")
                keyword = []
                self.dl(product_word_list)
                j = self.params.get("input_search_word_index", 0)
                while j < len(product_word_list):
                    self.dl(f"添加词语: {product_word_list[j]}")
                    if j >= len(product_word_list) - 1:
                        # 最后一个的情况
                        keyword.append(product_word_list[j])
                        self.dl("词语已经到结尾了")
                    elif len(keyword) < self.params.get("input_group_number", 5):
                        keyword.append(product_word_list[j])
                        j += 1
                        continue
                    self.dl(
                        f'词语范围: {j+1-self.params.get("input_group_number", 5)} - {j+1}')
                    self.dl(f"词语列表: {keyword}")
                    if not self.pi_liang_cha_ci(keyword):
                        self.close()
                        #  user_result = pyautogui.confirm(text="查词失败了,是否要继续呢",title="查词出错啦",buttons=["继续","跳过","x"])
                        user_result = True
                        if not user_result:
                            keyword = []
                            j += 1
                        continue
                    if not self.shuai_xuan():
                        self.close()
                        #  user_result = pyautogui.confirm(text="查词失败了,是否要继续呢",title="查词出错啦",buttons=["继续","跳过","x"])
                        user_result = True
                        if not user_result:
                            keyword = []
                            j += 1
                        continue
                    data = self.copy_all_data()
                    if data:
                        self.save_data(self.product_list[i][0], data)
                        self.add_black_list(data)
                        j += 1
                    else:
                        self.dl("获取数据失败")
                    self.close()
                    keyword = []
                    #  break
        except Exception as e:
            log.exception(e)
        # 功能执行结束
        self.exit()

    def main(self, params):
        self.init_var()
        self.params = params
        self.dl(params)
        self.read_product_list()
        fun = self.params.get("cmb_fun_switch")
        if fun == "复制标题":
            self.init_driver(12306)
            self.get_title_save_to_excel()
        elif fun == "批量查词":
            self.pi_liang_cha_ci_run()
        elif fun == "拆分词根":
            self.get_root_word_save_to_excel()
        self.set_status("脚本完成")


if __name__ == "__main__":
    print("自动化测试")
    run = ChuanQi()
    #  qq_login.main({
    #  "index": 0,
    #  "hwnd": 198880,
    #  })
    #  run.get_title_save_to_excel()
    #  run.read_keyword_list("置物架")
    #  root = run.get_root_word_list()
    #  run.save_root_word("置物架", root)
    #  run.save_title("测试", ["1", "2"])
    data = """关键词|搜索人数|搜索次数|点击率|点击人数|点击次数|交易金额|支付转化率|在线商品数|商城点击占比|直通车参考价|支付人数|客单价|搜索倍数|交易倍数|蓝海值
水槽置物架|1783|7390|81.47%|1066|6021|8160|7.5%|161773|88.06%|1.33|134|0.61|0.01|0.05|0.2
厨房置物架落地多层|1454|6841|94.07%|965|6435|18207|10.98%|172257|79.24%|1.49|160|1.14|0.01|0.11|0.49
不锈钢置物架|1447|6034|88.1%|827|5316|12932|8.95%|158004|37.1%|1.25|130|0.99|0.01|0.08|1.23
蔬菜置物架|1439|7014|100.67%|1001|7061|9529|8.19%|73926|83.98%|1.23|118|0.81|0.02|0.13|0.7
杯子收纳置物架|1404|6513|104.5%|886|6806|3886|6.09%|81854|63.06%|0.94|86|0.45|0.02|0.05|1.1
厨房调味品置物架|1400|6226|82.93%|817|5163|6138|9.79%|73863|88.42%|1.43|137|0.45|0.02|0.08|0.49"""
    result = run.get_data_list(data)
    print(result)
    run.save_data("查词", data)
    #  run.save_title("测试2", [])
