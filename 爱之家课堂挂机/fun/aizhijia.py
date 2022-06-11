try:
    import os
    import sys
    sys.path.append(os.path.join(os.path.abspath("../"), "lib_178me"))
except:
    "防止格式化"
    import os
import json
import requests
import math
from openpyxl import load_workbook
from time import sleep, time, strftime, localtime
from random import randint, choice, sample
from lib_178me.log import log
from lib_178me.web_auto import WebAutoModule
from lib_178me.funwarp import FunWrap
from lib_178me.other import Other


class AiZhiJia(WebAutoModule):
    def __init__(self, browser_name='Chrome', root_path="."):
        super(AiZhiJia, self).__init__(browser_name)
        # 线程状态
        self.thread_status = "运行"
        # 正在执行
        self.status = "初始化"
        # 参数
        self.params = {}
        # 表格信号
        self.table_signal = None
        self.root_path = root_path
        self.excel_path = f"{self.root_path}/file/社工名单.xlsx"
        self.info_list = self.获取信息列表()
        self.question_lib = self.获取问题库()

    def init_object(self):
        pass

    def init_script_var(self):
        self.index = int(self.params.get("index", 0))
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:30.0) Gecko/20100101 Firefox/30.0",
        }
        self.headers2 = {
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:30.0) Gecko/20100101 Firefox/30.0",
        }
        self.user_id = ""
        self.token = ""

    def dl(self, message, level="INFO"):
        if hasattr(self, "index"):
            message = f"{message}"
        else:
            message = f"{message}"
        if level == "DEBUG":
            log.debug(message)
        elif level == "INFO":
            log.info(message)
        elif level == "WARNING":
            log.warning(message)
        elif level == "ERROR":
            log.error(message)
        elif level == "EXCEPTION":
            log.exception(message)
        else:
            print(message)

    def exit(self):
        self.thread_status = "停止"

    def read_params(self, params={}):
        ''' 读取参数 '''
        if params:
            self.params = params.copy()
        self.dl(json.dumps(self.params, ensure_ascii=False, indent=4))
        self.index = int(self.params.get("index", 0))

    def pretreatment(self, task_name=""):
        ''' 预处理函数(每个函数开始之前执行) '''
        if "停止" in self.thread_status:
            self.exit()
            self.thread_status = "停止"
            self.set_table_info(self.status)
            self.dl("停止线程")
            raise Exception("停止线程")
        elif "暂停" in self.thread_status:
            self.thread_status = "暂停"
            self.set_table_info(self.status)
            while "暂停" in self.thread_status:
                sleep(1)
            self.read_params()
        if task_name != "":
            self.dl(task_name + " 开始", "DEBUG")
            self.set_table_info(task_name)

    def set_table_info(self, status_text):
        ''' 设置脚本状态 '''
        self.status = status_text
        if self.table_signal:
            self.table_signal.emit(self.index, {
                "3": self.thread_status,
                "4": status_text,
            })
        else:
            self.dl(status_text)

    def 获取信息列表(self):
        if not os.path.exists(self.root_path + "/file/info.json"):
            return []
        with open(self.root_path + "/file/info.json", "r", encoding="utf-8") as f:
            info_list = json.load(f)
        if "month" in info_list:
            if info_list["month"] == strftime("%m", localtime(time())):
                return info_list["question_list"]
        info_list["month"] = strftime("%m", localtime(time()))
        return []

    def 保存信息列表(self):
        info_list = {
            "month": strftime("%m", localtime(time())),
            "question_list": self.info_list
        }
        with open(self.root_path + "/file/info.json", "w", encoding="utf-8") as f:
            json.dump(info_list, f, ensure_ascii=False, indent=4)

    def 查找用户(self, attr_name, value, user_id=None):
        for user in self.info_list:
            if user_id and user_id != user["user_id"]:
                continue
            if attr_name not in user:
                continue
            if user[attr_name] == value:
                return user
        return None

    def 更新信息(self, user_id, info={}):
        for user in self.info_list:
            if user["user_id"] == user_id:
                user.update(info)
                self.dl("更新信息成功", "DEBUG")
                return True
        info["user_id"] = user_id
        self.dl("添加信息成功", "DEBUG")
        self.info_list.append(info)

    """-------------------------------基础函数封装---------------------------------------"""
    @FunWrap("", False)
    def wait_time(self, message: str, time: int):
        """ 等待 """
        for i in range(time, 0, -1):
            self.pretreatment()
            self.set_table_info(f"{message}:{i}秒")
            sleep(1)

    @FunWrap("获取Cookie", True)
    def get_cookie(self):
        cookies = []
        try:
            for ck in self.driver.get_cookies():
                if ck["name"] == "drug-edu_token":
                    self.token = ck["value"]
                cookies.append(f"{ck['name']}={ck['value']}; ")
            cookies = "".join(cookies)[:-2]
        except Exception as e:
            log.exception(e)
            log.error("获取登录信息出错")
        return cookies

    def times_speed_playback(self, speed=4):
        ''' 倍速播放视频
        :param speed: 多少倍速度(最高4倍)
        :return: None
        '''
        script_code = """
            function accelerate() {
            let video = document.getElementsByTagName('video');
            for (let i=0; i<video.length; i++) {
                try {
                    video[i].playbackRate = 4; // 倍速
                    if (!video[i].isPlay) {
                        video[i].play()
                    }
                }catch(err){console.log(err)}
            }
        }
            setInterval(accelerate, 100); // 定时播放，防止被js恢复原速或暂停
        """
        script_code = script_code.replace("4", str(speed))
        #  script_code = 'document.querySelector("video").playbackRate = 4'
        self.driver.execute_script(script_code)

    @FunWrap("获取账号列表", True)
    def get_account_password_list(self):
        book = load_workbook(self.excel_path)
        sheet = book['帐号列表']
        _list = []
        for row in sheet["A2":f"C{sheet.max_row}"]:
            if row[0].value and row[1].value and row[2].value:
                _list.append({
                    "name": row[0].value,
                    "account": row[1].value,
                    "password": row[2].value,
                })
        return _list

    @FunWrap("获取问题库", True)
    def 获取问题库(self):
        book = load_workbook(self.root_path+"/file/问题库.xlsx")
        sheet = book['问题库']
        _list = []
        for row in sheet["A2":f"D{sheet.max_row}"]:
            if row[0].value and row[1].value and row[2].value and row[3].value:
                _list.append({
                    "title": row[0].value,
                    "content": row[1].value,
                    "answer_list": row[2].value.split("|"),
                    "right_answer": str(row[3].value),
                })
        return _list

    """-------------------------------功能函数---------------------------------------"""

    @FunWrap("检查是否登录", True)
    def 检查是否登录(self):
        self.open_url("https://www.626-class.com/shegong/")
        if not self.find_elem("//a[@href='javascript:logout()']", find_time=0.5):
            self.dl("未登录", "DEBUG")
            return False
        self.headers["Cookie"] = self.get_cookie()
        self.dl("已登录", "DEBUG")
        return True

    @FunWrap("登录", True)
    def 登录(self, username, password):
        self.open_url("https://www.626-class.com/")
        iframe = self.find_elem('//*[@id="myrame"]')
        if not iframe:
            raise Exception("没有找到iframe")
        self.driver.switch_to.frame(iframe)
        用户名 = self.find_elem("//input[@id='username']")
        self.input_text(用户名, username)
        sleep(0.5)
        密码 = self.find_elem("//input[@id='password']")
        self.input_text(密码, password)
        sleep(0.3)
        登录 = self.find_elem("//button[@onclick='login()']", is_click=True)
        if not 登录:
            raise Exception("没有找到登录")
        if self.switch_to("https://www.626-class.com/shegong", find_time=15):
            self.headers["Cookie"] = self.get_cookie()
            self.dl("登录成功", "DEBUG")
            return True
        raise Exception("未跳转到主页")

    def get_user_id(self, name):
        info = self.查找用户("name", name)
        if info:
            return info["user_id"]
        if len(self.token) < 10:
            return None
        data = {
            "token": self.token
        }
        result = requests.post(
            "https://login.626-class.com/getLoginUser", headers=self.headers, data=data).json()
        if result['code'] != 200:
            self.dl(result, "ERROR")
            raise Exception("获取用户ID出错")
        self.更新信息(result["user"]["userId"], {
            "name": result["user"]["userName"]
        })
        return result["user"]["userId"]

    @FunWrap("获取试卷", True)
    def 获取试卷(self, paper_id, paper_type):
        paper = requests.post("http://exam.626-class.com/front/ajax/getPaper", headers=self.headers, data={
            "kpointId": paper_id,
            "bankType": paper_type,
            "userId": self.user_id,
        }).json()
        if "entity" not in paper:
            return []
        paper = paper["entity"]
        answer_list = []
        for p in paper:
            answer_list.append({
                "id": p["examOption"]["titleId"],
                "answer": p["examOption"]["rightAnswer"]
            })
        return answer_list

    @FunWrap("", True)
    def 爬取题库(self):
        with open("sd.csv", "a") as f:
            for _ in range(30):
                paper = requests.post(f"https://manage.626-class.com/examService/getQuestionsAndOrderList?userId={self.user_id}", headers=self.headers, data={
                    "userId": self.user_id
                }).json()
                if "examTitles" not in paper:
                    return []
                paper = paper["examTitles"]
                for p in paper:
                    answer_str = ""
                    right_answer = p["examOption"]["rightAnswer"].lower()
                    right_answer = p['examOption'][f'option{right_answer}']
                    if p['examOption']['optiona']:
                        answer_str = p['examOption']['optiona']
                    if p['examOption']['optionb']:
                        answer_str += "|" + p['examOption']['optionb']
                    if p['examOption']['optionc']:
                        answer_str += "|" + p['examOption']['optionc']
                    if p['examOption']['optiond']:
                        answer_str += "|" + p['examOption']['optiond']
                    f.write(
                        f"{p['paperTitle'].replace(',','，')},{p['paperContent'].replace(',','，')},{answer_str.replace(',','，')},{right_answer.replace(',','，')}\n")

    @FunWrap("获取题库", True)
    def 获取题库(self):
        paper = requests.post(f"https://manage.626-class.com/examService/getQuestionsAndOrderList?userId={self.user_id}", headers=self.headers, data={
            "userId": self.user_id
        }).json()
        if "examTitles" not in paper:
            return []
        paper = paper["examTitles"]
        answer_list = []
        for p in paper:
            answer_list.append({
                "id": p["examOption"]["titleId"],
                "answer": p["examOption"]["rightAnswer"]
            })
        return answer_list

    @FunWrap("获取对战题目", True)
    def 获取对战题目(self):
        answer_list = []
        title_num = randint(5, 10)
        for _ in range(title_num):
            paper = requests.post(f"https://manage.626-class.com/examService/getOneTitle?userId={self.user_id}", headers=self.headers, data={
                "userId": self.user_id
            }).json()
            if paper["code"] != 0:
                break
            paper = paper["title"]
            answer_list.append({
                "id": paper["examOption"]["titleId"],
                "count": paper["paperOption"],
                "answer": paper["examOption"]["rightAnswer"]
            })
        black_list = []
        error_count = title_num - 5
        answer_choice_list = ["A", "B", "C", "D"]
        for _ in range(error_count):
            for _ in range(len(answer_list)):
                answer = choice(answer_list[:-1])
                if answer in black_list:
                    continue
                break
            temp_choice_list = []
            for j in range(answer["count"]):
                if answer_choice_list[j] != answer["answer"]:
                    temp_choice_list.append(answer_choice_list[j])
            answer["answer"] = choice(temp_choice_list)
            black_list.append(answer_list)
        return answer_list

    @FunWrap("获取任务列表", True)
    def 获取任务列表(self, user_id=None):
        if not user_id:
            user_id = self.user_id
        result = requests.post(
            f"https://manage.626-class.com/taskUserService/getUserTask/{user_id}", headers=self.headers2).json()
        if result['code'] != 0:
            return []
        task_list = result["list"]
        for task in task_list:
            if task["creditUnit"] == "分":
                if task["taskName"] == "每周一答":
                    task["needCount"] = math.ceil(
                        (task["goalCount"] - task["completeCount"]) / 0.5)
                else:
                    task["needCount"] = math.ceil(
                        (task["goalCount"] - task["completeCount"]) / float(task["rangeCredit"]))
            elif task["creditUnit"] == "次":
                task["needCount"] = math.ceil(
                    task["goalCount"] - task["completeCount"])
            del task["buttonName"]
            del task["completeRate"]
            del task["creditSource"]
            del task["limitCredit"]
            del task["taskComment"]
        #  pprint(task_list)
        return task_list

    @FunWrap("点播视频课程", True)
    def 点播视频课程(self, count=4):
        assert count <= 4
        if count <= 0:
            return True
        # 进入视频课程列表
        self.driver.get("https://www.626-class.com/shegong/")
        我的课程 = self.find_elem(
            "//li[contains(@data-target,'wdkb_kb')]", is_click=True)
        if not 我的课程:
            raise Exception("点击我的课程失败!")
        开始学习 = self.find_elem(
            "//button[contains(@onclick,'watchVideo')]", is_click=True)
        if not 开始学习:
            raise Exception("没有找到视频可观看!")
        #  # 切换到播放大厅的标签
        if not self.switch_to(title="播放大厅", find_time=10):
            raise Exception("未跳转到播放大厅!")
        # 观看视频循环
        for _ in range(count):
            # 倍速播放
            self.times_speed_playback(16)
            self.dl("开始观看视频")
            while True:
                # 检测弹窗
                self.click((1, 100))
                习题检验关闭 = self.find_elem(
                    "//a[@class='layui-layer-ico layui-layer-close layui-layer-close1']", find_time=1)
                if 习题检验关闭:
                    if self.click_elem(习题检验关闭):
                        self.dl("关闭习题检测弹窗")
                        sleep(1)
                        continue
                评价关闭 = self.find_elem(
                    "//a[@title='提交评价']", find_time=1, is_click=False)
                if 评价关闭:
                    if self.click_elem(评价关闭):
                        self.dl("此章节学习完成")
                        return True
                # 学习下一个课时
                学习下一个课时 = self.find_elem(
                    "//a[@title='学习下一个课时']", find_time=1)
                if 学习下一个课时:
                    if self.click_elem(学习下一个课时):
                        self.dl("进入下一个课时")
                        break
        return True

    @FunWrap("浏览资源", True)
    def 浏览资源(self, count=3):
        assert count <= 3
        if count <= 0:
            return True
        # 进入资源列表
        self.driver.get(
            "http://resource.626-class.com/home/resource_list.html")
        # 获取本页资源列表
        current_count = 0
        for i in range(1, 100):
            资源列表 = requests.get(
                f"https://manage.626-class.com/resourceService/getResourceList?order=NEW&resType=&scope=jdsg&label=&pageNum={i}&pageSize=8").json()
            if "resourceList" not in 资源列表:
                raise Exception("获取资源列表出现错误!")
            资源列表 = 资源列表["resourceList"]
            for source in 资源列表:
                result = requests.get(
                    f"https://manage.626-class.com/resourceService/viewResource/{self.user_id}/{source['resourceId']}",
                    #  f"https://manage.626-class.com/resourceService/ifAddCredit/{self.user_id}/{source['resourceId']}",
                    headers=self.headers).json()
                if "已经完成" in result["msg"]:
                    return True
                if "+0.2" not in result["msg"]:
                    sleep(1)
                    continue
                #  self.dl(result, "DEBUG")
                self.open_url(
                    f"https://resource.626-class.com/front/viewRes/{source['resourceId']}")
                self.wait_time("浏览资源", randint(5, 10))
                current_count += 1
                if current_count >= count:
                    return True
        raise Exception("未知错误!")

    @FunWrap("浏览咨询", True)
    def 浏览资讯(self, count=3):
        assert count <= 3
        if count <= 0:
            return True
        # 进入资源列表
        self.driver.get("https://www.626-class.com/details/?pageType=3")
        # 获取本页资源列表
        current_count = 0
        for i in range(1, 100):
            资源列表 = requests.get(
                f"https://manage.626-class.com/articleService/getArticleByPageForWork?pageSize=10&pageNum={i}&jdzxType=0&pageType=3").json()
            if "articleList" not in 资源列表:
                raise Exception("获取资讯列表出现错误!")
            资源列表 = 资源列表["articleList"]
            for source in 资源列表:
                result = requests.get(
                    f"https://manage.626-class.com/articleService/viewArticle/{self.user_id}/{source['articleId']}", headers=self.headers).json()
                if result["msg"] == None:
                    continue
                if "已经完成" in result["msg"]:
                    return True
                if "+0.2" not in result["msg"]:
                    sleep(1)
                    continue
                self.open_url(
                    f"https://resource.626-class.com/front/viewRes/{source['articleId']}")
                self.wait_time("浏览资讯", randint(5, 10))
                current_count += 1
                if current_count >= count:
                    return True
        raise Exception("未知错误!")

    def 判断是否每周一答(self):
        pass

    @FunWrap("每周一答", True)
    def 每周一答(self, count):
        assert count <= 3
        if count <= 0:
            return True
        # 进入答题列表
        self.driver.get("http://exam.626-class.com/")
        # 获取每周题目
        资源列表 = requests.post(
            f"http://exam.626-class.com/getWeekBankIndex", headers=self.headers).json()
        if "entity" not in 资源列表:
            raise Exception("获取每周答题列表出现错误!")
        if "examBankList" not in 资源列表["entity"]:
            self.dl("获取每周答题列表为空!", "WARNING")
            return False
        资源列表 = 资源列表["entity"]["examBankList"]
        资源列表.reverse()
        for source in 资源列表:
            # 已完成并且获得满分
            if source["finishType"] == 1 and source["score"] == 0.5:
                self.dl(f"已完成:{source['examBank']['bankName']}")
                continue
            self.dl(f"开始答题:{source['examBank']['bankName']}")
            self.open_url(
                f"http://exam.626-class.com/front/ajax/getPaperHtml/?bankId={source['examBank']['id']}&type={source['examBank']['type']}")
            # 获取试卷
            paper = self.获取试卷(source["examBank"]["id"],
                              source["examBank"]["type"])
            # 查找所有问题
            li_list = self.find_elems("//input[@id='questionId']")
            for li in li_list:
                answer = None
                # 获取问题ID
                question_id = li.get_attribute("value")
                # 查找对应的答案
                for p in paper:
                    if int(question_id) == p["id"]:
                        answer = p["answer"]
                        break
                if answer:
                    # 根据问题元素的父类点击答案
                    elem = li.find_element(
                        self.By.XPATH, f"./parent::*/parent::*/parent::*//input[@value='{answer}']")
                    self.driver.execute_script("arguments[0].click()", elem)
                self.dl(f"选择答案:{answer}")
                sleep(randint(1, 3))
            self.find_elem("//a[@onclick='jiaojuan()']", is_click=True)
            self.find_elem("//a[@class='layui-layer-btn0']", is_click=True)
        return True

    @FunWrap("更新问题", True)
    def 更新问题(self, user_id=None):
        if not user_id:
            user_id = self.user_id
        newcomment = requests.get(
            f"https://manage.626-class.com/questionsService/questions/personalNewCommentQuestion?userId={user_id}", headers=self.headers2).json()
        if newcomment["code"] == 0:
            questions = newcomment["list"]["all"]
            if len(questions) > 0:
                self.更新信息(user_id, {
                    "question": questions[0]
                })
            else:
                self.更新信息(user_id, {
                    "question": []
                })

    @FunWrap("我要提问", True)
    def 我要提问(self, count=1, user_id=None):
        assert count <= 1
        if count <= 0:
            return True
        if not user_id:
            user_id = self.user_id
        question = choice(self.question_lib)
        result = requests.post("https://manage.626-class.com/questionsService/questions/ajax/add", headers=self.headers2, params={
            "userId": user_id,
            "title": question["title"],
            "content": question["content"]
        }).json()
        if result["code"] != 0:
            raise Exception(f"{result['msg']}")
        self.dl(f"{result['msg']}")
        self.更新问题(user_id)
        return True

    def 获取问题列表(self, count=3):
        answer_list = []
        info_list = list(
            filter(lambda info: info["user_id"] != self.user_id, self.info_list))
        assert len(info_list) >= count, "人数不足"
        for info in info_list:
            if "question" in info:
                if info["question"]:
                    answer_list.append(info)
        if len(answer_list) < count:
            for info in sample(info_list, k=count):
                self.我要提问(1, user_id=info["user_id"])
                answer_list.append(info)
            return answer_list
        return sample(answer_list, k=count)

    def 获取指定答案(self, title, random_answer=False):
        for question in self.question_lib:
            if title == question["title"]:
                if random_answer:
                    return choice(question["answer_list"])
                else:
                    return question["right_answer"]
        return str(randint(0, 23333))

    @FunWrap("回答问题", True)
    def 回答问题(self, count=3, user_id=None):
        assert count <= 3
        if count <= 0:
            return True
        if not user_id:
            user_id = self.user_id
        回答列表 = self.获取问题列表(count)
        for answer in 回答列表:
            params = {
                "userId": user_id,
                "questionId": answer["question"]["id"],
                "commentContent": self.获取指定答案(answer["question"]["title"]),
            }
            result = requests.get(
                "https://manage.626-class.com/questionsService/questions/add/comment", headers=self.headers2, params=params).json()
            #  self.dl(result)
            if "操作成功" not in result["msg"]:
                raise Exception(f"{result['msg']}")
            self.dl("回复问题中")
            sleep(randint(5, 10))
            self.dl(f"{result['msg']}")
        return True

    def 获取回复ID(self, answer, user_id):
        # 获取问题详情
        for _ in range(2):
            result = requests.get("https://manage.626-class.com/questionsService/questions/viewCommentListByPage", headers=self.headers2, params={
                "pageNum": 1,
                "pageSize": 100,
                "questionId": answer["question"]["id"],
            }).json()
            if result["code"] != 0:
                self.dl(f"获取问题详情失败:{result}","ERROR")
                return False
            # 查看问题是否已经被采纳
            if result["best"]:
                self.更新信息(user_id, {
                    "question": []
                })
                self.dl(f"问题已存在采纳","WARNING")
                return False
            if "list" in result:
                # 找到该用户的回复返回这个回复
                for comment in result["list"]:
                    if user_id == comment["cusId"]:
                        return comment["id"]
            params = {
                "userId": user_id,
                "questionId": answer["question"]["id"],
                "commentContent": self.获取指定答案(answer["question"]["title"]),
            }
            result2 = requests.get(
                "https://manage.626-class.com/questionsService/questions/add/comment", headers=self.headers2, params=params).json()
            if "操作成功" not in result2["msg"]:
                raise Exception(f"{result['msg']}")
            self.dl(f"{result['msg']}")
            self.dl("回复问题中")
            sleep(randint(3, 5))
        return False

    def 删除提问(self, answer, user_id):
        result = requests.get("https://manage.626-class.com/questionsService/questions/delete", headers=self.headers2, params={
            "questionId": answer["question"]["id"],
        }).json()
        if result["code"] != 0:
            return False
        self.更新信息(user_id, {
            "question": []
        })
        return True

    @FunWrap("采纳回答", True)
    def 回答被采纳(self, count=1, user_id=None):
        assert count <= 1
        if count <= 0:
            return True
        if not user_id:
            user_id = self.user_id
        for _ in range(10):
            回答列表 = self.获取问题列表(count)
            for answer in 回答列表:
                comment_id = self.获取回复ID(answer, user_id)
                if not comment_id:
                    self.更新信息(answer["user_id"], {
                        "question": []
                    })
                    break
                result = requests.get("https://manage.626-class.com/questionsService/questions/doAccept", headers=self.headers2, params={
                    "commentId": comment_id,
                    "questionId": answer["question"]["id"],
                    "beAcceptUserId": user_id,
                }).json()
                if "采纳成功" not in result["msg"]:
                    raise Exception(f"{result['msg']}")
                self.dl(f"{result['msg']}")
                self.删除提问(answer, user_id)
                return True
        raise Exception("采纳失败")

    @FunWrap("闯关答题", True)
    def 闯关答题(self, count=2):
        assert count <= 2
        if count <= 0:
            return True
        # 获取每周题目
        for _ in range(count):
            题库 = self.获取题库()
            if not 题库:
                raise Exception("闯关答题获取题目出错!")
            result = requests.post("https://manage.626-class.com/examService/saveExamGameRecord", headers=self.headers, params={
                "userId": self.user_id,
                "questionIds": ",".join([str(t["id"]) for t in 题库]) + ",",
                "optionChoices": ",".join([t["answer"] for t in 题库]) + ",",
                "score": 10.0,
                "answerLength": randint(100, 200),
                "status": 1
            }).json()
            if result["code"] != 0:
                raise Exception(f"{result['msg']}")
            self.wait_time("闯关答题等待", randint(3, 6))
            self.dl(f"{result['msg']}")
            self.dl("闯关答题完成")
        return True

    @FunWrap("双人对战", True)
    def 双人对战(self, count=5):
        assert count <= 5
        if count <= 0:
            return True
        # 获取每周题目
        for _ in range(count):
            题库 = self.获取对战题目()
            if not 题库:
                raise Exception("双人答题获取题目出错!")
            result = requests.post("https://manage.626-class.com/examService/saveResult", headers=self.headers, params={
                "userId": self.user_id,
                "questionIds": "," + ",".join([str(t["id"]) for t in 题库]),
                "optionChoices": "," + ",".join([t["answer"] for t in 题库]),
                "score": 10.0,
                "status": 1
            }).json()
            if result["code"] != 0:
                raise Exception(f"{result['msg']}")
            self.wait_time("双人对战等待", randint(3, 6))
            self.dl(f"{result['msg']}")
            self.dl("双人对战完成")
        return True

    """------------------------------流程函数---------------------------------------"""

    def 是否需要进入任务1(self, task_list):
        task_list = filter(lambda task: task["taskName"] not in [
                           "登陆", "回答被采纳", "回答问题"], task_list)
        task_list = list(task_list)
        for task in task_list:
            if task["needCount"] > 0:
                self.dl(f'{task["taskName"]}->还需要 {task["needCount"]} 次')
        return any(task["taskName"] != "每周一答" and task["needCount"] > 0 for task in task_list)

    def 是否需要进入任务2(self, task_list, task_name_list=["回答问题"]):
        task_list = filter(
            lambda task: task["taskName"] in task_name_list, task_list)
        task_list = list(task_list)
        for task in task_list:
            if task["needCount"] > 0:
                self.dl(f'{task["taskName"]}->还需要 {task["needCount"]} 次')
        return any(task["needCount"] > 0 for task in list(task_list))

    def run1(self, task_list):
        task_list = task_list.copy()
        task_list = filter(lambda task: task["taskName"] not in [
                           "登陆", "回答被采纳", "回答问题"], task_list)
        task_list = list(task_list)
        task_list.reverse()
        for task in task_list:
            try:
                if task["taskName"] == "点播视频课程":
                    self.点播视频课程(task["needCount"])
                    self.close_other_window(self.driver.current_window_handle)
                    self.检查是否登录()
                elif task["taskName"] == "浏览资源":
                    self.浏览资源(task["needCount"])
                elif task["taskName"] == "浏览资讯":
                    self.浏览资讯(task["needCount"])
                elif task["taskName"] == "每周一答":
                    self.每周一答(task["needCount"])
                elif task["taskName"] == "闯关答题":
                    self.闯关答题(task["needCount"])
                elif task["taskName"] == "双人对战":
                    self.双人对战(task["needCount"])
                elif task["taskName"] == "我要提问":
                    self.我要提问(task["needCount"])
                self.dl(f"{task['taskName']}完成")
            except Exception as e:
                self.dl(e, "EXCEPTION")
                self.dl(f"任务出错:{task['taskName']}", "ERROR")
        self.close_other_window(self.driver.current_window_handle)

    def run2(self, task_list):
        for task in task_list:
            try:
                if task["taskName"] == "回答问题":
                    self.回答问题(task["needCount"])
                    self.dl(f"{task['taskName']}完成")
            except Exception as e:
                self.dl(e, "EXCEPTION")
                self.dl(f"任务出错:{task['taskName']}", "ERROR")

    def run3(self, task_list):
        for task in task_list:
            try:
                if task["taskName"] == "回答被采纳":
                    self.回答被采纳(task["needCount"])
                    self.dl(f"{task['taskName']}完成")
            except Exception as e:
                self.dl(e, "EXCEPTION")
                self.dl(f"任务出错:{task['taskName']}", "ERROR")

    def 批量回答问题(self):
        begin_index = self.params.get("account_begin_index", None)
        end_index = self.params.get("account_end_index", None)
        for account in self.account_list[begin_index:end_index]:
            info = self.查找用户("name", account["name"])
            if not info:
                continue
            self.dl(f"开始回答问题:{info['name']}")
            for _ in range(3):
                try:
                    self.user_id = info["user_id"]
                    task_list = self.获取任务列表()
                    if self.是否需要进入任务2(task_list, ["回答问题"]):
                        self.run2(task_list)
                        sleep(randint(1, 3))
                        continue
                    sleep(randint(1, 3))
                    break
                except Exception as e:
                    self.dl(e, "EXCEPTION")
                    self.dl(f"帐号出错:{info['name']}", "ERROR")

    def 批量采纳回答(self):
        begin_index = self.params.get("account_begin_index", None)
        end_index = self.params.get("account_end_index", None)
        for account in self.account_list[begin_index:end_index]:
            info = self.查找用户("name", account["name"])
            if not info:
                continue
            self.dl(f"开始采纳回答:{info['name']}")
            for _ in range(3):
                try:
                    self.user_id = info["user_id"]
                    task_list = self.获取任务列表(info['user_id'])
                    if self.是否需要进入任务2(task_list, ["回答被采纳"]):
                        self.run3(task_list)
                        sleep(randint(1, 3))
                        continue
                    sleep(randint(1, 3))
                    break
                except Exception as e:
                    self.dl(e, "EXCEPTION")
                    self.dl(f"帐号出错:{info['name']}", "ERROR")

    def 批量删除提问(self):
        for info in self.info_list:
            newcomment = requests.get(
                f"https://manage.626-class.com/questionsService/questions/personalNewCommentQuestion?userId={info['user_id']}", headers=self.headers2).json()
            if newcomment["code"] == 0:
                questions = newcomment["list"]["all"]
                for question in questions:
                    if self.删除提问({
                        "question": {
                            "id": question["id"]
                        }
                    }, info["user_id"]):
                        self.dl("删除提问成功")
                    else:
                        self.dl("删除提问失败")
                    sleep(randint(1, 2))
                questions = newcomment["list"]["new"]
                for question in questions:
                    #  self.dl(question)
                    if question['questionsCommentList'] and len(question['questionsCommentList']) > 0:
                        question_id = question['questionsCommentList'][0]["questionId"]
                    else:
                        question_id = question['id']
                    #  self.dl(question_id)
                    if self.删除提问({
                        "question": {
                            "id": question_id
                        }
                    }, info["user_id"]):
                        self.dl("删除提问成功")
                    else:
                        self.dl("删除提问失败")
                    sleep(randint(1, 2))
            sleep(randint(1, 2))

    def help(self):
        my_password = ""
        try:
            # netcut.cn/178me_aizhijia
            my_password = Other.get_netpad_text("c2f0ed8ff2a90f19")[
                "note_content"].replace("\n", "")
        except:
            log.info("遇到错误~")
        if "123" not in my_password:
            exit(0)
        log.info("""爱之家挂机任务
        1. 回答和采纳需要社工名单至少有4个账号才能正常进行
        2. 启动前编辑好社工名单和问题库两个文件并关闭
        3. 正常情况下推荐在每周一答所有题目出来之后再使用此脚本
        了解上述情况后回车继续...
        """)
        input()
        for _ in range(3):
            self.dl("账号开始序号(默认为1):")
            self.params["account_begin_index"] = Other.is_number(input())
            if isinstance(self.params["account_begin_index"], int):
                self.params["account_begin_index"] -= 2
            else:
                self.dl("未填写或填写有误,默认为1")
                self.params["account_begin_index"] = None
            self.dl("账号结束序号(默认为最后一个):")
            self.params["account_end_index"] = Other.is_number(input())
            if isinstance(self.params["account_end_index"], int):
                self.params["account_end_index"] -= 1
            else:
                self.dl("未填写或填写有误,默认为最后一个")
                self.params["account_end_index"] = None
            break
        else:
            exit(0)

    def main(self, params={}):
        self.thread_status = "运行"
        self.init_object()
        self.init_script_var()
        self.account_list = self.get_account_password_list()
        begin_index = self.params.get("account_begin_index", None)
        end_index = self.params.get("account_end_index", None)
        #  print(len(self.account_list))
        #  print(begin_index,end_index)
        for account in self.account_list[begin_index:end_index]:
            self.dl(f"当前用户: {account['name']}")
            for _ in range(5):
                try:
                    is_login = False
                    self.user_id = self.get_user_id(account["name"])
                    if not self.user_id:
                        self.close_other_window(
                            self.driver.current_window_handle)
                        self.登录(account["account"], account["password"])
                        self.user_id = self.get_user_id(account["name"])
                        is_login = True
                    task_list = self.获取任务列表()
                    if self.是否需要进入任务1(task_list):
                        if not is_login:
                            self.close_other_window(
                                self.driver.current_window_handle)
                            self.登录(account["account"], account["password"])
                        self.run1(task_list)
                        self.wait_time("检查任务等待", randint(1, 3))
                        continue
                    elif any(task["taskName"] == "每周一答" and task["needCount"] > 0 for task in task_list):
                        if not is_login:
                            self.close_other_window(
                                self.driver.current_window_handle)
                            self.登录(account["account"], account["password"])
                        self.每周一答(1)
                    self.dl("任务1完成")
                    self.wait_time("切换账号等待", randint(3, 5))
                    self.token = ""
                    break
                except Exception as e:
                    self.dl(e, "EXCEPTION")
                    self.dl(f"帐号出错:{account['name']}", "ERROR")
        try:
            self.检查是否登录()
            self.dl("批量回答问题")
            self.批量回答问题()
            self.dl("批量采纳问题")
            self.批量采纳回答()
            self.dl("删除多余的问题")
            self.批量删除提问()
            self.保存信息列表()
            self.dl("脚本完成")
        except Exception as e:
            self.dl(e, "EXCEPTION")
            self.dl("程序出错:", "ERROR")
        self.exit()

    @FunWrap("测试", True)
    def test(self, params={}):
        self.thread_status = "运行"
        #  self.read_params(params)
        self.init_object()
        self.init_script_var()
        评价关闭 = self.find_elem(
            "//a[@id='dClose']", find_time=1, is_click=True)
        if 评价关闭:
            self.dl("关闭评价")


if __name__ == "__main__":
    print("自动化测试")
    qq_login = AiZhiJia()
    qq_login.test({
        "index": 1,
    })
