import datetime
import os
import openpyxl
import requests
import logging
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

#  excel 操作 可视化
#  Request 爬虫 信息
#  selenium 网页

class Hisense:
    def __init__(self):
        print("开始初始化")
        self.read_params()
        self.read_keyword_list()
        self.read_location_list()
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Mobile Safari/537.36",
            "Cookie": self.params.get("cookies", "")
        }
        self.sid = self.params.get("sid", "EAueKlCCfkcsBnbnfPCCRzwaULtGOFzy")
        self.sheet_name = "-1"
        self.mail_list = []
        self.session = requests.session()
        self.driver = None

    def init_driver(self, port=None, browser_name="360安全浏览器"):
        options = Options()
        if port:
            options.add_experimental_option(
                "debuggerAddress", f"127.0.0.1:{port}")
        if browser_name != "":
            browser_path = self.get_browser_path(browser_name)
            options.binary_location = browser_path
            self.driver = webdriver.Chrome(
                "./input/chromedriver.exe", options=options)
        else:
            service = webdriver.chrome.service.Service("./chromedriver")
            self.driver = webdriver.Chrome(service=service,options=options)
        print(f"连接{browser_name}成功!")

    def get_browser_path(self, browser_name):
        import win32com.client
        try:
            shell = win32com.client.Dispatch("WScript.Shell")
            path = os.path.join(os.path.expanduser(r'~'), "Desktop")
            for dirpath, dirnames, filenames in os.walk(path):
                for filename in filenames:
                    if browser_name in filename:
                        path = dirpath + "\\" + filename
            shortcut = shell.CreateShortCut(path).TargetPath
            return shortcut
        except Exception as e:
            print(e)
            raise Exception(f"{browser_name}路径查找失败")

    def set_cookies(self):
        if not self.driver:
            self.init_driver(12306)
        hisense_cookies = self.driver.get_cookies()
        cookies = ""
        for ck in hisense_cookies:
            if ck["name"] in ["Coremail.sid"]:
                self.sid = ck["value"]
            elif ck["name"] in ["ssoLoginToken", "Coremail", "Coremail.sid"]:
                cookies += f"{ck['name']}={ck['value']}; "
        cookies = cookies[:-2]
        self.headers["Cookie"] = cookies

    def login(self):
        # 检查cookies 是否可用 正常继续 不正常则获取cookies
        result = None
        for _ in range(2):
            try:
                method = "getAllFolders"
                url = f'https://mail.hisense.com/coremail/XT5/jsp/mail.jsp?func={method}&sid={self.sid}'
                result = self.session.get(url, headers=self.headers)
                result = result.json()
                if "S_OK" not in result["code"]:
                    raise Exception("登陆信息已失效")
                print("验证成功!")
                self.write_params()
                return True
            except Exception as e:
                print(e)
                self.set_cookies()
        return False

    def read_keyword_list(self):
        print("读取关键词列表")
        self.keyword_list = []
        book = openpyxl.load_workbook("./out/邮件.xlsx")  # 打开exce文件
        sheet = book["关键词列表"]
        for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=4):
            keyword = []
            for cell in row:
                keyword.append(cell.value)
            self.keyword_list.append(keyword)

    def read_location_list(self):
        print("读取邮件位置对应表")
        self.location_list = []
        book = openpyxl.load_workbook("./out/邮件.xlsx")  # 打开exce文件
        sheet = book["邮件位置对应表"]
        for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=2):
            keyword = []
            for cell in row:
                keyword.append(cell.value)
            self.location_list.append(keyword)

    def read_params(self):
        print("读取脚本参数")
        book = openpyxl.load_workbook("./out/邮件.xlsx")  # 打开exce文件
        sheet = book["脚本参数"]
        self.params = {
            "cookies": sheet["A2"].value,
            "sid": sheet["B2"].value,
        }

    def write_params(self):
        print("保存脚本参数")
        book = openpyxl.load_workbook("./out/邮件.xlsx")  # 打开exce文件
        sheet = book["脚本参数"]
        sheet["A2"].value = self.headers["Cookie"]
        sheet["B2"].value = self.sid
        book.save("./out/邮件.xlsx")

    def write_location_list(self, location_list):
        print("保存邮件位置对应表")
        book = openpyxl.load_workbook("./out/邮件.xlsx")
        sheet = book["邮件位置对应表"]
        for i, location in enumerate(location_list):
            for j, value in enumerate(location):
                if i == 0 and j == 0:
                    sheet.cell(row=i+1, column=j+1).value = "代码"
                elif i == 0 and j == 1:
                    sheet.cell(row=i+1, column=j+1).value = "位置"
                else:
                    sheet.cell(row=i+1, column=j+1).value = value
        book.save("./out/邮件.xlsx")

    def search_message(self, date_rang):
        print("搜索指定邮件")
        data = {
            "start": (None, 0),
            "limit": (None, 100),
            "fid": (None, 0),  # 文件夹
            #  "sentDate": (None, "2022-01-15:2022-01-16"),  # 时间
            "sentDate": (None, date_rang),  # 时间
        }
        #  with open("./data.json", "r") as f:
        #  return json.load(f)
        method = "searchMessages"
        url = f'https://mail.hisense.com/coremail/XT5/jsp/mail.jsp?func={method}&sid={self.sid}'
        result = self.session.post(url, data=data, headers=self.headers)
        result = result.json()
        if "S_OK" not in result["code"]:
            raise Exception(f"请求失败:{result['code']}")
        #  print(result)
        return result["var"]

    def get_attach_list(self, mid):
        print("读取邮件附件地址")
        data = {
            "mid": (None, mid)
        }
        method = "readMessage"
        url = f'https://mail.hisense.com/coremail/XT5/jsp/mail.jsp?func={method}&sid={self.sid}'
        result = self.session.post(url, data=data, headers=self.headers)
        result = result.json()
        if "S_OK" not in result["code"]:
            raise Exception(f"请求失败:{result['code']}")
        attach_list = []
        for attach in result["var"]["mail"]["attachments"]:
            if "filename" in attach:
                if "image" in attach["contentType"]:
                    continue
                attach_list.append({
                    "filename": attach["filename"],
                    #  "url": f"https://mail.hisense.com/coremail/mbox-data/附件：{attach['filename']}?part={attach['id']}&mode=preview&mboxa=&mid={mid}"
                    "url": f"https://mail.hisense.com/coremail/common/preview/preview.jsp?part={attach['id']}&mode=preview&mboxa=&mid={mid}"
                })
        return attach_list

    def get_mail_addr(self, fid, mid):
        #  return f'https://mail2.hisense.com/coremail/XT5/index.jsp?sid={self.sid}#mail.read|{{"fid":"{fid}","mid":"{mid}","mboxa":"","start":0}}'
        return f'https://mail.hisense.com/coremail/hxphone/sso.html#/frame/read/{fid}/{mid}'

    def get_attach_addr(self, mid):
        return self.get_attach_list(mid)

    def get_location(self, fid):
        for _ in range(2):
            for location in self.location_list:
                if fid == location[0]:
                    return location[1]
            method = "getAllFolders"
            url = f'https://mail.hisense.com/coremail/XT5/jsp/mail.jsp?func={method}&sid={self.sid}'
            result = self.session.get(url, headers=self.headers)
            result = result.json()
            if "S_OK" not in result["code"]:
                print("请求失败!")
                continue
            location_list = [(value["id"], value["name"])
                             for value in result["var"]]
            self.write_location_list(location_list)
            self.read_location_list()
        return ""

    def get_keyword(self, title):
        for keyword in self.keyword_list:
            if keyword[3] in title:
                return keyword[0:-1]
        return ["", "", ""]

    def get_mail_info(self, date_rang):
        date = datetime.datetime.strptime(date_rang, "%Y-%m-%d")
        date_rang = date_rang + ":" + \
            (date + datetime.timedelta(days=1)).strftime("%Y-%m-%d")
        mail_list = self.search_message(date_rang)
        print("筛选邮件信息...")
        for mail in mail_list:
            info = {}
            info["location"] = self.get_location(mail["fid"])
            info["title"] = mail["subject"]
            info["date"] = mail["sentDate"]
            info["from"] = mail["from"]
            info["mail_addr"] = self.get_mail_addr(mail["fid"], mail["id"])
            attach_addr = self.get_attach_addr(mail["id"])
            info["attach_addr"] = attach_addr
            info["attach_name"] = attach_addr
            keyword = self.get_keyword(mail["subject"])
            info["keyword1"] = keyword[0]
            info["keyword2"] = keyword[1]
            info["keyword3"] = keyword[2]
            self.mail_list.append(info)

    def set_excel(self):
        self.sheet_name = self.mail_list[-1]["date"]
        self.sheet_name = self.sheet_name.replace("-", "")[0:8]
        book = openpyxl.load_workbook("./out/邮件.xlsx")  # 打开exce文件
        for _sheet_name in book.sheetnames:
            if "heet" in _sheet_name:
                _sheet_name = book.get_sheet_by_name(_sheet_name)
                book.remove(_sheet_name)
        if self.sheet_name not in book.sheetnames:
            book.create_sheet(self.sheet_name)
        sheet = book[self.sheet_name]  # 打开工作表
        sheet.column_dimensions['A'].width = 13
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 18
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['G'].width = 30
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 15
        sheet.column_dimensions['J'].width = 15
        sheet.cell(1, 1).value = "邮件位置"
        sheet.cell(1, 2).value = "邮件标题"
        sheet.cell(1, 3).value = "邮件时间"
        sheet.cell(1, 4).value = "发件人"
        sheet.cell(1, 5).value = "邮件地址"
        sheet.cell(1, 6).value = "附件预览"
        sheet.cell(1, 7).value = "附件预览地址"
        sheet.cell(1, 8).value = "关键词1（分类）"
        sheet.cell(1, 9).value = "关键词2（对象）"
        sheet.cell(1, 10).value = "关键词3（事项）"
        book.save("./out/邮件.xlsx")

    def write_excel(self):
        print(f"保存{self.sheet_name}工作表")
        book = openpyxl.load_workbook("./out/邮件.xlsx")
        sheet = book[self.sheet_name]
        for info in self.mail_list:
            row = sheet.max_row + 1
            for j, key in enumerate(info):
                if key in ["attach_addr","attach_name"]:
                    attach_row = row
                    for attach in info[key]:
                        if key in ["attach_name"]:
                            sheet.cell(row=attach_row, column=j +
                                       1).value = attach["url"]
                        else:
                            sheet.cell(row=attach_row, column=j +
                                       1).value = attach["filename"]
                            sheet.cell(row=attach_row, column=j +
                                       1).hyperlink = attach["url"]
                        attach_row += 1
                    continue
                sheet.cell(row=row, column=j+1).value = info[key]
                if "mail_addr" in key:
                    sheet.cell(row=row, column=j+1).value = info[key]
                    sheet.cell(row=row, column=j+1).hyperlink = info[key]
        book.save("./out/邮件.xlsx")

    def main(self):
        try:
            if not self.login():
                input("获取登录信息失败~")
                exit(0)
            while True:
                try:
                    print("欢迎使用邮箱查询工具(Ctrl-c 强制停止!)")
                    date_rang = input("输入指定日期查询 例如(2022-01-15)\n")
                    self.get_mail_info(date_rang)
                    self.set_excel()
                    self.write_excel()
                    if input("回车继续使用,q Q退出脚本! --> ") in ['q', "Q"]:
                        print("脚本已退出!")
                        break
                    os.system("cls")
                except Exception as e:
                    print("脚本出现以下错误")
                    logging.exception(e)
                    input("尝试重新运行或者联系作者~")
        except Exception as e:
            print("脚本出现以下错误")
            logging.exception(e)
            input("尝试重新运行或者联系作者~")


if __name__ == "__main__":
    h = Hisense()
    h.main()
